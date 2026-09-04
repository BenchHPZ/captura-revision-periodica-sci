#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae "CHECK LIST A-01 2026.pdf" (RAG 4.1 -- checklist diario de la
ambulancia) a un Excel de trabajo humano. El Excel NO lo lee la
aplicación: es para que el área de PCI revise y corrija antes de
convertirlo (a mano, con ayuda de una IA) al JSON que importará /rag.

A diferencia de scripts/extraer_rags.py, aquí el catálogo de renglones
NO se infiere por regex: se comprobó a mano que el texto que pypdf
extrae de este PDF no conserva el orden visual de la tabla (columnas y
banners de categoría intercalados fuera de orden). El catálogo de ~115
renglones se transcribió leyendo cada página, mismo criterio que
PLANTILLAS en extraer_rags.py para datos que no son extraíbles con
confianza por su cuenta.

Lo que sí se automatiza:
  - extracción de las imágenes de referencia incrustadas por página
  - volcado del catálogo transcrito a las hojas del Excel

Señala ambigüedades del documento de origen -- Pos duplicados, huecos de
numeración, banner de categoría que no cambia entre páginas -- en la
hoja "Anomalias", en vez de resolverlas por su cuenta (mismo principio
que extraer_rags.py, ver docs/decisiones.md D-03).

Uso:
    python scripts/extraer_checklist.py \
        --pdf "C:\\Users\\brive\\Downloads\\CHECK LIST A-01 2026.pdf"

Por defecto escribe dentro de extracciones/RAG-4.1_Ambulancia-A01/ (ver
--salida / --salida-imagenes para cambiarlo). Esa carpeta no se versiona
-- ver .gitignore -- porque contiene datos operativos de la empresa.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

import pypdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


class Anomalias:
    """Recolecta avisos para la hoja 'Anomalias'; nunca interrumpe la extracción."""

    def __init__(self) -> None:
        self.items: list[tuple[str, str]] = []

    def registrar(self, contexto: str, mensaje: str) -> None:
        self.items.append((contexto, mensaje))

    def imprimir(self) -> None:
        print()
        if not self.items:
            print("Sin anomalías detectadas.")
            return
        print(f"{len(self.items)} anomalía(s) -- revisar en la hoja 'Anomalias' antes de convertir a JSON:")
        for contexto, mensaje in self.items:
            print(f"  - [{contexto}] {mensaje}")


# --------------------------------------------------------------- identidad

FORMATO = {
    "clave": "RAG 4.1",
    "nombre": 'Lista de Inspección de Ambulancia Planta Motores, Silao, GTO. "Urgencias Básicas"',
    "documento_referencia": "I1.15M2_4037-004",
    "revision": None,
    "notas": (
        "El PDF de origen es de 2022-2023, orientación vertical y llenado a mano en papel; el nuevo "
        "formato estandarizado se propone horizontal (ver plan de ampliación de RAGs). El encabezado "
        "trae AÑO/MES/Fecha/Grupo/Reviso Nombre repetido en cada página del checklist de equipo."
    ),
}

# ------------------------------------------------------------ portada: fotos

PORTADA_FOTOS = [
    {"orden": 1, "etiqueta": "Frontal", "archivo_foto_sugerido": None, "notas": None},
    {"orden": 2, "etiqueta": "Lateral / trasera derecha", "archivo_foto_sugerido": None, "notas": None},
    {"orden": 3, "etiqueta": "Trasera", "archivo_foto_sugerido": None, "notas": None},
    {"orden": 4, "etiqueta": "Lateral / trasera izquierda", "archivo_foto_sugerido": None, "notas": None},
]

# ---------------------------------------------------- checklist de equipo


def _item(categoria: str, pos: str | None, nombre: str, cantidad: str | None,
          verificaciones: list[str], notas: str | None = None) -> dict:
    return {
        "categoria": categoria,
        "pos_original": pos,
        "nombre_equipo": nombre,
        "cantidad": cantidad,
        "verificaciones": verificaciones,
        "archivo_foto_sugerido": None,
        "notas_ambiguedad": notas,
    }


# 'categoria' es el banner gris SECUNDARIO más reciente visto antes de
# este ítem (Equipo médico / Insumos / Soluciones y medicamentos /
# Botiquín de ambulancia / Extra) -- no el banner superior de página, que
# repite "RECURSOS FISICOS DE APOYO" en casi todas las páginas incluso
# cuando el contenido es claramente de otra categoría (ver anomalía
# "banner-persistente" más abajo).
CHECKLIST_EQUIPO = [
    # RECURSOS FISICOS DE APOYO -- página 2
    _item("RECURSOS FISICOS DE APOYO", "1", "Cinturones de seguridad en todos los asientos", "6",
          ["Buen estado", "Cantidad", "Limpieza"]),
    _item("RECURSOS FISICOS DE APOYO", "2", "Herramienta de mano", "1", ["Buen estado"]),
    _item("RECURSOS FISICOS DE APOYO", "3", "Señalización", "1", ["Buen estado"]),
    _item("RECURSOS FISICOS DE APOYO", "4", "Juego de cables pasa corriente", "1", ["Buen estado"]),
    _item("RECURSOS FISICOS DE APOYO", "5", "Lámpara portátil de emergencia", "1", ["Buen estado", "Cargadores"]),
    _item("RECURSOS FISICOS DE APOYO", "6", "Neumático de refacción con accesorios", "1",
          ["Gato, llave de cruz", "Neumático"]),
    _item("RECURSOS FISICOS DE APOYO", "7", "Extintor ABC", "1", ["Buen estado"]),
    _item("RECURSOS FISICOS DE APOYO", "8", "Equipo de comunicación", None, ["Buen estado", "Funcionando"]),
    # EQUIPO MEDICO -- banner gris nuevo, página 2
    _item("EQUIPO MEDICO", "8", "Camilla rígida con sistema de sujeción", "1",
          ["Araña", "Buen estado y limpio"],
          "Pos '8' repetido (mismo número que Equipo de comunicación, categoría distinta)."),
    _item("EQUIPO MEDICO", "9", "Carro Camilla", "1", ["Buen estado", "Limpieza", "Funcionalidad"]),
    _item("EQUIPO MEDICO", "10", "BVM", "1", ["Lactantes", "Pediátrico", "Adulto"]),
    # página 3 -- banner superior sigue diciendo RECURSOS FISICOS DE
    # APOYO pero el contenido es evidentemente EQUIPO MEDICO sin nuevo
    # banner gris secundario; se preserva EQUIPO MEDICO por continuidad.
    _item("EQUIPO MEDICO", "13", "Equipo de aspiración de secreciones fijo", "1", ["Buen estado y funcionando"]),
    _item("EQUIPO MEDICO", "14", "Equipo de aspiración de secreciones portátil", "1", ["Buen estado"]),
    _item("EQUIPO MEDICO", "16", "Gancho porta suero doble", "2", ["Buen estado"]),
    _item("EQUIPO MEDICO", None, "Collarines rígidos, adulto y pediátrico", "1", ["Buen estado", "Limpios"],
          "Sin número de Pos visible en el PDF."),
    _item("EQUIPO MEDICO", None, "Cánulas nasofaríngeas", "5", ["Buen estado", "Vigentes"],
          "Sin número de Pos visible en el PDF."),
    _item("EQUIPO MEDICO", None, "Dispositivo para inmovilizar la cabeza (bloques)", "1",
          ["Buen estado", "Limpios"], "Sin número de Pos visible en el PDF."),
    _item("EQUIPO MEDICO", None, "Desfibrilador automatizado externo", "1", ["Buen estado", "Con batería"],
          "Sin número de Pos visible en el PDF."),
    _item("EQUIPO MEDICO", None, "Dispositivo de estabilización pélvica", "1", ["Buen estado"],
          "Sin número de Pos visible en el PDF."),
    _item("EQUIPO MEDICO", "17",
          "Kit de parto; dos pinzas rochester, onfalotomo, tijeras tipo mayo, cinta umbilical, "
          "perilla y campos desechables", "1", ["Buen estado", "Completo"]),
    _item("EQUIPO MEDICO", "19", "Férulas para miembros torácicos y pélvicos", "1", ["Buen estado"],
          "Pos '19' se reutiliza más adelante para 'Tanque de oxígeno portátil tipo D'."),
    _item("EQUIPO MEDICO", "21", "Sistema de inmovilización pediátrica", "1", ["Buen estado"]),
    # página 4 -- banner superior RECURSOS FISICOS DE APOYO, contenido
    # todavía EQUIPO MEDICO/O2, luego banner gris nuevo INSUMOS.
    _item("EQUIPO MEDICO", "22", "Chaleco de extracción", "1", ["Buen estado"],
          "Pos '22' se reutiliza más abajo para 'Apositos estériles' (categoría INSUMOS)."),
    _item("EQUIPO MEDICO", "19", "Tanque de oxígeno portátil tipo D", "1",
          ["Regulador, llave y manómetro", "Mínimo de 1000 psi"],
          "Pos '19' repetido (ver Férulas arriba)."),
    _item("EQUIPO MEDICO", "20", "Tanque de oxígeno fijo tipo M", "1",
          ["Contenido mínimo 30%", "Regulador y manómetro", "Flujómetro y humificador"]),
    # INSUMOS -- banner gris nuevo, página 4
    _item("INSUMOS", "22", "Apositos estériles", "4", ["Buen estado y completas", "Vigentes"],
          "Pos '22' repetido (ver Chaleco de extracción arriba)."),
    _item("INSUMOS", "23", "Gasas estériles", "25", ["Buen estado y completas", "Vigentes"]),
    _item("INSUMOS", "24", "Cómodo", "1", ["Buen estado"]),
    _item("INSUMOS", "25", "Punzos", "1 C/U", ["No.14", "No.16", "No.18", "No.20", "No.21", "No.22"]),
    _item("INSUMOS", "26", "Contenedor rígido RPBI", "1", ["Buen estado"]),
    _item("INSUMOS", "27", "Desinfectante", "1", ["Buen estado"]),
    # página 5 -- banner superior RECURSOS FISICOS DE APOYO; sigue
    # INSUMOS sin nuevo banner hasta que aparece SOLUCIONES Y
    # MEDICAMENTOS en la página 7.
    _item("INSUMOS", "28", "Desinfectante para equipos y superficies", "1", ["Buen estado y completas"]),
    _item("INSUMOS", "29", "Normogotero y microgotero", "1 C/U", ["Buen estado", "Vigentes"]),
    _item("INSUMOS", "30", "Guantes", "10", ["Buen estado"]),
    _item("INSUMOS", "31", "Cubrebocas", "5", ["Buen estado"]),
    _item("INSUMOS", "32", "Jabón quirúrgico Benzal e isodine", "1", ["Contenido arriba del 50%"]),
    _item("INSUMOS", "33", "Jeringas", "1 pza/c.u.", ["1 ml", "3 ml", "5 ml", "10 ml", "20 ml"]),
    _item("INSUMOS", "34", "Ligadura", "1", ["Buen estado"]),
    _item("INSUMOS", "35", "Pato orinal", "1", ["Buen estado"]),
    _item("INSUMOS", "36", "Puntas nasales, adulto y pediátrico", "1 C/U", ["Buen estado", "Vigencia"]),
    _item("INSUMOS", "37", "Mascarilla simple, adulto y pediátrico", "1 C/U", ["Buen estado", "Vigencia"]),
    _item("INSUMOS", "38", "Mascarilla reservorio, adulto y pediátrico", "1 C/U", ["Buen estado", "Vigencia"]),
    # página 6 -- banner superior RECURSOS FISICOS DE APOYO, sin banner
    # gris secundario visible; contenido sigue pareciendo INSUMOS.
    _item("INSUMOS", "39", "Bolsa amarilla y roja para RPBI", "2 C/U", ["Buen estado", "Completas"]),
    _item("INSUMOS", "40", "Cánula de Yankauer", "1", ["Buen estado", "Vigencia"]),
    _item("INSUMOS", "41", "Guía para identificación de materiales peligrosos", "1", ["Buen estado"]),
    _item("INSUMOS", "42", "Sábana térmica", "1", ["Buen estado"]),
    _item("INSUMOS", "43", "Sábana para quemados", "1", ["Buen estado"]),
    _item("INSUMOS", "39", "Riñón", "1", ["Buen estado"], "Pos '39' repetido (ver Bolsa RPBI arriba)."),
    _item("INSUMOS", "40", "Cobertor", "1", ["Buen estado y limpio"],
          "Pos '40' repetido (ver Cánula de Yankauer arriba)."),
    _item("INSUMOS", "41", "Sondas de aspiración suave", "1 C/U", ["5 Fr", "14 Fr", "12 Fr", "8 Fr", "10 Fr"],
          "Pos '41' repetido (ver Guía para materiales peligrosos arriba)."),
    _item("INSUMOS", "42/43", "Tela adhesiva", "2", ["Buen estado"],
          "El PDF muestra '42' y '43' fusionados en una sola celda de Pos para este renglón."),
    _item("INSUMOS", "44", "Torundas secas", None, ["Buen estado"]),
    _item("INSUMOS", "45", "Torundas con alcohol", "15", ["Buen estado"]),
    # página 7 -- banner superior RECURSOS FISICOS DE APOYO; banner gris
    # nuevo SOLUCIONES Y MEDICAMENTOS a mitad de página.
    _item("INSUMOS", "46", "Vendas", "2 C/U", ["5 cm", "10 cm", "20 cm", "30 cm"]),
    # SOLUCIONES Y MEDICAMENTOS -- banner gris nuevo, página 7
    _item("SOLUCIONES Y MEDICAMENTOS", "53", "Cloruro de sodio 0.9%", "1 C/U", ["Vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "54", "Glucosa 5%", None, ["Vigente"],
          "Comparte la cantidad '1 C/U' con Cloruro de sodio/Solución hartman/Dextrosa (celda combinada)."),
    _item("SOLUCIONES Y MEDICAMENTOS", "55", "Solución hartman", None, ["Vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "56", "Dextrosa al 50%", None, ["Vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "57", "Ácido acetilsalicílico", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "58", "Isosorbida", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "59", "Trinitrato de glicerilo", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "60", "Adrenalina", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "61", "Atropina", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "62", "Epinefrina", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Ácido tranexámico", "1", ["Buen estado y vigente"],
          "Pos '63' se repite consecutivamente para varios medicamentos distintos en el PDF de origen."),
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Dopamina", "1", ["Buen estado y vigente"], "Pos '63' repetido."),
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Adenosina", "1", ["Buen estado y vigente"], "Pos '63' repetido."),
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Amiodarona", "1", ["Buen estado y vigente"],
          "Pos '63' repetido; este renglón queda justo antes del salto de página 7\u21927 -- "
          "confirmar que no es el mismo renglón duplicado por la paginación (ver el siguiente ítem)."),
    # página 8 -- continúa SOLUCIONES Y MEDICAMENTOS sin nuevo banner
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Amiodarona", "1", ["Buen estado y vigente"],
          "Aparece 'Amiodarona' de nuevo justo tras el bloque Nombre/Firma de fin de página anterior -- "
          "posible duplicado por paginación, no un segundo medicamento real. Confirmar con el área."),
    _item("SOLUCIONES Y MEDICAMENTOS", "63", "Salbutamol aerosol", "1", ["Buen estado y vigente"],
          "Pos '63' repetido."),
    _item("SOLUCIONES Y MEDICAMENTOS", "64", "Electrolitos orales", "2", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "65", "Glucosa vía oral", "1", ["Buen estado y vigente"]),
    _item("SOLUCIONES Y MEDICAMENTOS", "66", "Agua inyectable", "1", ["Buen estado y vigente"]),
    # BOTIQUIN DE AMBULANCIA -- banner gris nuevo, página 8
    _item("BOTIQUIN DE AMBULANCIA", "67", "Glucómetro", "1",
          ["Batería con carga", "No. lancetas", "No. tiras reactivas"]),
    _item("BOTIQUIN DE AMBULANCIA", "68", "Cánulas orofaríngeas", "7", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "69", "Lámpara pupilera", "1", ["Buen estado", "Batería con carga"]),
    _item("BOTIQUIN DE AMBULANCIA", "70", "Estetoscopio pinard", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "71", "Abatelenguas de plástico", "1", ["Buen estado"]),
    # página 9 -- banner superior RECURSOS FISICOS DE APOYO, sin banner
    # gris secundario visible; contenido sigue BOTIQUIN DE AMBULANCIA,
    # luego banner gris nuevo EXTRA a mitad de página.
    _item("BOTIQUIN DE AMBULANCIA", "72", "Oxímetro", "1", ["Buen estado", "Batería con carga"]),
    _item("BOTIQUIN DE AMBULANCIA", "73", "Termómetro", "1 C/U", ["Digital", "Pistola"]),
    _item("BOTIQUIN DE AMBULANCIA", "74", "Tijera de uso rudo", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "75", "Escudo facial para RCP", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "76", "Apositos estériles", "5", ["Buen estado"],
          "Nombre repetido de 'Apositos estériles' (Pos 22, categoría INSUMOS) -- cantidad distinta."),
    _item("BOTIQUIN DE AMBULANCIA", "77", "Gasas estériles", "20", ["Buen estado", "Cantidad"],
          "Nombre repetido de 'Gasas estériles' (Pos 23, categoría INSUMOS) -- cantidad distinta."),
    _item("BOTIQUIN DE AMBULANCIA", "78", "Vendas", "2 C/U", ["5 cm", "10 cm", "20 cm", "30 cm"],
          "Nombre repetido de 'Vendas' (Pos 46, categoría INSUMOS)."),
    _item("BOTIQUIN DE AMBULANCIA", "79", "Normogotero/microgotero", "1 C/U", ["Buen estado y sellado"]),
    _item("BOTIQUIN DE AMBULANCIA", "80", "Toallas alcoholadas", "20", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "81", "Tegaderm", "2", ["Buen estado", "Cantidad"]),
    _item("BOTIQUIN DE AMBULANCIA", "82", "Ligadura", "1", ["Buen estado"]),
    # página 10 -- banner superior RECURSOS FISICOS DE APOYO, contenido
    # sigue BOTIQUIN DE AMBULANCIA sin nuevo banner.
    _item("BOTIQUIN DE AMBULANCIA", "83", "Tela adhesiva", "2", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "84", "Corta anillos", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "85", "Baumanómetro", "1", ["Adulto", "Pediátrico"]),
    _item("BOTIQUIN DE AMBULANCIA", "86", "Estetoscopio", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "87", "Sábana térmica", "1", ["Buen estado"],
          "Nombre repetido de 'Sábana térmica' (Pos 42, categoría INSUMOS)."),
    _item("BOTIQUIN DE AMBULANCIA", "89", "Jabón quirúrgico e isodine", "1 C/U",
          ["Recipiente en buen estado", "Contenido mayor al 50%"],
          "No hay Pos '88' en el PDF -- salta de 87 a 89."),
    _item("BOTIQUIN DE AMBULANCIA", "90", "Torniquete", "2", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "91", "Jeringas", "1 pza/c.u.", ["1 ml", "3 ml", "5 ml", "10 ml"],
          "Nombre repetido de 'Jeringas' (Pos 33, categoría INSUMOS) -- aquí sin la presentación de 20 ml."),
    _item("BOTIQUIN DE AMBULANCIA", "92", "Punzos", "2 C/U",
          ["No.14", "No.16", "No.18", "No.20", "No.21", "No.22"],
          "Nombre repetido de 'Punzos' (Pos 25, categoría INSUMOS) -- cantidad distinta."),
    _item("BOTIQUIN DE AMBULANCIA", "93", "Rastrillo", "2", ["Buen estado"]),
    # página 11 -- banner superior RECURSOS FISICOS DE APOYO, contenido
    # sigue BOTIQUIN DE AMBULANCIA, luego banner gris nuevo EXTRA.
    _item("BOTIQUIN DE AMBULANCIA", "94", "Extractor de venenos", "1 kit", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "95", "Bolsa válvula mascarilla, adulto", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "96", "Mascarilla con reservorio, adulto", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "97", "Tarjetas triage", "25", ["Buen estado", "Cantidad"]),
    _item("BOTIQUIN DE AMBULANCIA", "98", "Puntas nasales, adulto", "1", ["Buen estado"],
          "Nombre parecido a 'Puntas nasales, adulto y pediátrico' (Pos 36, categoría INSUMOS)."),
    _item("BOTIQUIN DE AMBULANCIA", "99", "Lava ojos", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "100", "Mascarilla pocket", "1", ["Buen estado"]),
    _item("BOTIQUIN DE AMBULANCIA", "101", "Guantes estériles", "5 pares", ["Buen estado"]),
    # EXTRA -- banner gris nuevo, página 11
    _item("EXTRA", "102", "Doppler", "1", ["Buen estado", "Con batería"]),
    _item("EXTRA", "103", "Monitor", "1", ["Buen estado", "Derivadas, electrodos y oxímetro", "Funcionando"]),
    _item("EXTRA", "104", "Kit de diagnóstico", "3", ["Buen estado"]),
    _item("EXTRA", "105", "Camilla marina", "1", ["Buen estado"]),
    # página 12 -- banner superior RECURSOS FISICOS DE APOYO, contenido
    # sigue EXTRA sin nuevo banner; termina el checklist de equipo.
    _item("EXTRA", "106", "Estuche de disección", "1", ["Buen estado"]),
    _item("EXTRA", "107", "Férula moldeable", "1", ["Buen estado"]),
    _item("EXTRA", "108", "Carbón activado", "1", ["Buen estado"]),
    _item("EXTRA", "109", "Perilla de aspiración", "1", ["Buen estado"]),
]

for _indice, _renglon in enumerate(CHECKLIST_EQUIPO, start=1):
    _renglon["orden"] = _indice

# ------------------------------------------------- sub-checklist mecánico

_MECANICO_POS_DESCRIPCION = [
    (1, "Nivel adecuado de aceite"),
    (2, "Nivel adecuado de anticongelante"),
    (3, "Nivel adecuado líquido de frenos"),
    (4, "Batería en buen estado"),
    (5, "Combustible (MAL<50%>BIEN)"),
    (6, "Sirena"),
    (7, "Luces (Baja/Alta)"),
    (8, "Cuartos"),
    (9, "Direccionales"),
    (10, "Estacionarias"),
    (11, "Limpiadores"),
    (12, "Claxon"),
    (13, "Luz interior"),
    (14, "Luz y alarma de reversa"),
    (15, "Luz stop"),
    (16, "Torreta"),
    (17, "Estrobos"),
    (18, "Luces laterales (blancas)"),
    (19, "Luces laterales y traseras intermitentes"),
    (25, "Seguro vigente"),
    (26, "Tarjeta de circulación"),
    (27, "Kilometraje"),
]

CHECKLIST_MECANICO = [
    {
        "pos_original": str(pos),
        "orden": orden,
        "descripcion": descripcion,
        "notas_ambiguedad": "No hay Pos 20 a 24 en el PDF -- salta de 19 a 25." if pos == 25 else None,
    }
    for orden, (pos, descripcion) in enumerate(_MECANICO_POS_DESCRIPCION, start=1)
]

# --------------------------------------------------------- bitácora de insumos

BITACORA_COLUMNAS = [
    "Insumo utilizado o faltante",
    "Cantidad",
    "Motivo",
    "Folio de FRAP o ID de práctica",
    "Nombre de quien reporta",
    "Fecha de reposición",
]
BITACORA_FILAS_BLANCO = 11  # filas en blanco visibles en la página de bitácora del PDF


def extraer_imagenes(pdf_path: Path, salida: Path) -> int:
    reader = pypdf.PdfReader(str(pdf_path))
    salida.mkdir(parents=True, exist_ok=True)
    total = 0
    for num_pagina, pagina in enumerate(reader.pages, start=1):
        for i, img in enumerate(pagina.images, start=1):
            sufijo = Path(img.name).suffix or ".png"
            (salida / f"pagina-{num_pagina:02d}-img-{i:02d}{sufijo}").write_bytes(img.data)
            total += 1
    return total


def construir_anomalias() -> Anomalias:
    anomalias = Anomalias()

    contador_pos = Counter(it["pos_original"] for it in CHECKLIST_EQUIPO if it["pos_original"])
    for pos, veces in sorted(contador_pos.items()):
        if veces > 1:
            anomalias.registrar("Checklist_Equipo", f"Pos '{pos}' aparece {veces} veces")

    sin_pos = sum(1 for it in CHECKLIST_EQUIPO if not it["pos_original"])
    if sin_pos:
        anomalias.registrar("Checklist_Equipo", f"{sin_pos} renglón(es) sin número de Pos visible en el PDF")

    anomalias.registrar(
        "Checklist_Equipo",
        "el banner gris superior de página dice 'RECURSOS FISICOS DE APOYO' de forma persistente en "
        "casi todas las páginas 3-12 aunque el contenido sea de otra categoría; se usó el último banner "
        "gris SECUNDARIO visto (el que sí cambia a media página) como 'categoria' de cada renglón -- "
        "confirmar con el área si esto es correcto o si el banner superior debía haberse actualizado.",
    )
    anomalias.registrar(
        "Checklist_Mecanico",
        "no hay Pos 20 a 24 en el sub-checklist mecánico del vehículo -- salta de 19 (luces laterales "
        "y traseras intermitentes) a 25 (seguro vigente).",
    )
    return anomalias


def _formatear_encabezado(ws) -> None:
    relleno = PatternFill("solid", fgColor="002733")
    fuente = Font(color="FFFFFF", bold=True)
    for celda in ws[1]:
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _hoja_formato(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Formato"
    ws.append(["clave", "nombre", "documento_referencia", "revision", "notas"])
    ws.append([FORMATO["clave"], FORMATO["nombre"], FORMATO["documento_referencia"],
               FORMATO["revision"], FORMATO["notas"]])
    _formatear_encabezado(ws)
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["E"].width = 70


def _hoja_portada_fotos(wb: Workbook) -> None:
    ws = wb.create_sheet("Portada_Fotos")
    ws.append(["orden", "etiqueta", "archivo_foto_sugerido", "notas"])
    for f in PORTADA_FOTOS:
        ws.append([f["orden"], f["etiqueta"], f["archivo_foto_sugerido"], f["notas"]])
    _formatear_encabezado(ws)
    ws.column_dimensions["B"].width = 30


def _hoja_checklist_equipo(wb: Workbook) -> None:
    ws = wb.create_sheet("Checklist_Equipo")
    ws.append(["categoria", "pos_original", "orden", "nombre_equipo", "cantidad",
               "verificaciones", "archivo_foto_sugerido", "notas_ambiguedad"])
    for it in CHECKLIST_EQUIPO:
        ws.append([it["categoria"], it["pos_original"], it["orden"], it["nombre_equipo"], it["cantidad"],
                   "; ".join(it["verificaciones"]), it["archivo_foto_sugerido"], it["notas_ambiguedad"]])
    _formatear_encabezado(ws)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["H"].width = 60


def _hoja_checklist_mecanico(wb: Workbook) -> None:
    ws = wb.create_sheet("Checklist_Mecanico")
    ws.append(["pos_original", "orden", "descripcion", "notas_ambiguedad"])
    for it in CHECKLIST_MECANICO:
        ws.append([it["pos_original"], it["orden"], it["descripcion"], it["notas_ambiguedad"]])
    _formatear_encabezado(ws)
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 50


def _hoja_bitacora(wb: Workbook) -> None:
    ws = wb.create_sheet("Bitacora_Config")
    ws.append(["columna", "orden", "filas_blanco"])
    for i, columna in enumerate(BITACORA_COLUMNAS, start=1):
        ws.append([columna, i, BITACORA_FILAS_BLANCO if i == 1 else None])
    _formatear_encabezado(ws)
    ws.column_dimensions["A"].width = 32


def _hoja_anomalias(wb: Workbook, anomalias: Anomalias) -> None:
    ws = wb.create_sheet("Anomalias")
    ws.append(["hoja", "fila_referencia", "descripcion", "decision_sugerida"])
    for contexto, mensaje in anomalias.items:
        ws.append([contexto, None, mensaje, None])
    _formatear_encabezado(ws)
    ws.column_dimensions["C"].width = 95
    ws.column_dimensions["D"].width = 35


def main() -> int:
    carpeta_unidad = Path("extracciones") / "RAG-4.1_Ambulancia-A01"

    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--pdf", required=True, help="Ruta al PDF de origen (CHECK LIST A-01 2026.pdf)")
    ap.add_argument("--salida", default=str(carpeta_unidad / "checklist_A01_extraccion.xlsx"))
    ap.add_argument("--salida-imagenes", default=str(carpeta_unidad / "imagenes"))
    args = ap.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.is_file():
        print(f"No existe el archivo: {pdf_path}", file=sys.stderr)
        return 1

    salida_imagenes = Path(args.salida_imagenes)
    total_imagenes = extraer_imagenes(pdf_path, salida_imagenes)
    print(f"Imágenes extraídas: {total_imagenes} -> {salida_imagenes}/")

    anomalias = construir_anomalias()

    wb = Workbook()
    _hoja_formato(wb)
    _hoja_portada_fotos(wb)
    _hoja_checklist_equipo(wb)
    _hoja_checklist_mecanico(wb)
    _hoja_bitacora(wb)
    _hoja_anomalias(wb, anomalias)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)

    print(f"\nExcel escrito: {salida}")
    print(f"  Formato: 1 fila")
    print(f"  Portada_Fotos: {len(PORTADA_FOTOS)} filas")
    print(f"  Checklist_Equipo: {len(CHECKLIST_EQUIPO)} filas")
    print(f"  Checklist_Mecanico: {len(CHECKLIST_MECANICO)} filas")
    print(f"  Bitacora_Config: {len(BITACORA_COLUMNAS)} filas")
    anomalias.imprimir()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
